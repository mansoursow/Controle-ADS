import io
import os
import re
import sys
import json
import glob as glob_module
import base64
import time
import traceback
import unicodedata
import hashlib
from pathlib import Path
import pandas as pd
import openai
from openai import AsyncOpenAI
from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from dotenv import load_dotenv

load_dotenv()

app = FastAPI(title="AccounTech AI - Audit Expert Multi-Gares")

ALLOWED_ORIGIN = os.getenv("ALLOWED_ORIGIN", "http://localhost:5174")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[ALLOWED_ORIGIN],
    allow_methods=["POST"],
    allow_headers=["*"],
)

client = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Tarifs système ouvert (fixe par gare, peu importe l'origine)
TARIFS_FIXES = {
    "THIAMBOKH": {"C1": 500,  "C2": 1000, "C3": 1500, "C4": 2000},
    "AIBD":      {"C1": 500,  "C2": 1000, "C3": 1500, "C4": 2000},
}

# Tarifs système fermé Ila Touba (bidirectionnel)
TARIFS_FERMES = {
    ("BAMBEY",   "DIOURBEL"): {"C1": 500,  "C2": 500,  "C3": 1000, "C4": 1500},
    ("BAMBEY",   "THIES"):    {"C1": 1000, "C2": 1500, "C3": 2500, "C4": 3500},
    ("BAMBEY",   "TOUBA"):    {"C1": 1000, "C2": 1000, "C3": 2000, "C4": 3000},
    ("DIOURBEL", "THIES"):    {"C1": 1500, "C2": 2500, "C3": 3500, "C4": 5000},
    ("DIOURBEL", "TOUBA"):    {"C1": 500,  "C2": 1000, "C3": 1500, "C4": 2000},
    ("THIES",    "TOUBA"):    {"C1": 2000, "C2": 3000, "C3": 4500, "C4": 6500},
}

ALGO_PROMPT = """Tu es un auditeur de péage routier au Sénégal. Classifie le véhicule visible sur la photo selon l'Arrêté interministériel n°014313 du 26 avril 2019.

CATÉGORIES OFFICIELLES :

C1 — MOTO uniquement
- Motocyclettes, scooters, motos seules

C2 — VÉHICULE PARTICULIER et TAXI
- Berlines, citadines, breaks, coupés
- SUV, 4x4, tout-terrain (Prado, Pajero, Land Cruiser, RAV4…)
- Pick-up avec benne/plateau (HiLux, L200, Navara, Amarok…)
- Taxis

C3 — MINIBUS et CAMIONNETTE
- Fourgonnettes et vans utilitaires (Transit, Sprinter, Trafic…)
- Minibus (capacité ≤ 30 places)

C4 — BUS et POIDS LOURD (CAMION)
- Camions, semi-remorques, porteurs, bennes
- Bus de grande capacité (> 30 places), autocars
- Engins de chantier et véhicules de travaux

RÈGLE ABSOLUE : en cas de doute entre deux catégories, choisis la plus élevée.

Réponds UNIQUEMENT par un code parmi : C1, C2, C3, C4"""


def calculer_tarif(classe_ia: str, gare_entree: str, gare_sortie: str) -> int:
    entree = gare_entree.strip().upper()
    sortie = gare_sortie.strip().upper()

    for keyword, tarifs in TARIFS_FIXES.items():
        if keyword in entree or keyword in sortie:
            return tarifs.get(classe_ia, 0)

    MOTS_CLES = ["BAMBEY", "DIOURBEL", "THIES", "TOUBA"]
    kw_e = next((k for k in MOTS_CLES if k in entree), None)
    kw_s = next((k for k in MOTS_CLES if k in sortie), None)
    if kw_e and kw_s and kw_e != kw_s:
        cle = tuple(sorted([kw_e, kw_s]))
        if cle in TARIFS_FERMES:
            return TARIFS_FERMES[cle].get(classe_ia, 0)

    return 0


# Erreur fatale qui stoppe tout l'audit
class QuotaEpuiseeError(Exception):
    pass


async def analyze_transaction(tx_id, paye_agent, img_binary, gare_entree, gare_sortie,
                               classe_facturee="", classe_sortie="", date_sortie=""):
    result = {
        "id": tx_id,
        "date_sortie": date_sortie,
        "classe_agent": f"{paye_agent} FCFA",
        "classe_facturee": classe_facturee,
        "classe_sortie": classe_sortie,
        "gare_entree": gare_entree,
        "gare_sortie": gare_sortie,
        "categorie_ia": "N/A",
        "prix_theorique": "N/A",
        "ecart": 0,
        "image_data": None
    }
    if not img_binary:
        return result
    try:
        base64_img = base64.b64encode(img_binary).decode("utf-8")
        result["image_data"] = base64_img

        response = await client.chat.completions.create(
            model="gpt-4o-mini",
            max_tokens=10,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_img}",
                                "detail": "low"
                            }
                        },
                        {
                            "type": "text",
                            "text": ALGO_PROMPT + "\n\nClasse ce véhicule :"
                        }
                    ],
                }
            ],
        )

        texte = response.choices[0].message.content.strip().upper()
        # Extraire C1/C2/C3/C4 de la réponse
        classe_ia = "N/A"
        for code in ["C4", "C3", "C2", "C1"]:
            if code in texte:
                classe_ia = code
                break

        result["categorie_ia"] = classe_ia
        if classe_ia != "N/A":
            prix_du = calculer_tarif(classe_ia, gare_entree, gare_sortie)
            result["prix_theorique"] = f"{prix_du} FCFA"
            result["ecart"] = int(paye_agent) - prix_du

        return result

    except openai.RateLimitError as e:
        print(f"[IA][QUOTA] {tx_id} — quota épuisé : {e}", flush=True)
        raise QuotaEpuiseeError(str(e))
    except openai.APIStatusError as e:
        if e.status_code in (402, 429):
            raise QuotaEpuiseeError(str(e))
        print(f"[IA][ERREUR] {tx_id} — {type(e).__name__}: {e}", flush=True)
        return result
    except Exception as e:
        print(f"[IA][ERREUR] {tx_id} — {type(e).__name__}: {e}", flush=True)
        traceback.print_exc()
        return result


def get_photo_key(row):
    photo_col = row.get('PhotoSortie', '')
    if photo_col and str(photo_col).strip() not in ('', 'nan'):
        return os.path.splitext(os.path.basename(str(photo_col)))[0].strip()
    return str(row.get('NumeroSortie', row.get('NumeroTransaction', ''))).strip()


@app.post("/audit")
async def audit_csv(request: Request):
    form_data = await request.form(max_files=50_000, max_fields=50_000)
    file = form_data.get("file")
    images = [v for k, v in form_data.multi_items() if k == "images"]
    is_last_chunk = form_data.get("is_last_chunk", "true") == "true"
    chunk_index   = int(form_data.get("chunk_index", "0"))

    async def generate():
        t_start = time.time()
        try:
            # --- 1. Lecture CSV ---
            content = await file.read()
            df = pd.read_csv(io.BytesIO(content), sep=None, engine='python')
            df.columns = [c.strip() for c in df.columns]
            print(f"[AUDIT] Chunk {chunk_index} | CSV: {len(df)} lignes | Images: {len(images)}", flush=True)

            # --- 2. Map images ---
            images_map = {
                os.path.splitext(os.path.basename(img.filename))[0].strip(): img
                for img in images
            }

            # --- 3. Pré-matching ---
            rows_with_img    = []
            rows_without_img = []
            for _, row in df.iterrows():
                key = get_photo_key(row)
                if key in images_map:
                    rows_with_img.append((key, row))
                else:
                    rows_without_img.append(row)

            total_with    = len(rows_with_img)
            total_without = len(rows_without_img)
            print(f"[AUDIT] Chunk {chunk_index} | Avec image: {total_with} | Sans image: {total_without}", flush=True)
            yield f"data: {json.dumps({'type': 'info', 'total_with_img': total_with, 'total_without_img': total_without})}\n\n"

            # --- 4. Analyse IA ---
            for idx, (img_key, row) in enumerate(rows_with_img):
                if await request.is_disconnected():
                    print(f"[AUDIT] Client déconnecté. Arrêt.", flush=True)
                    return

                tx_id = str(row.get('NumeroSortie', row.get('NumeroTransaction', ''))).strip()
                elapsed = time.time() - t_start
                print(f"[AUDIT] Chunk {chunk_index} [{idx+1}/{total_with}] ID={tx_id} | {elapsed:.1f}s", flush=True)

                paye_val = row.get('TotalPaye', row.get('Espece', 0))
                try:
                    paye = int(float(str(paye_val).replace(',', '.')))
                except (ValueError, TypeError):
                    paye = 0

                g_entree        = str(row.get('NomGareEntree', 'THIAMBOKH')).strip()
                g_sortie        = str(row.get('NomGareSortie', 'THIAMBOKH')).strip()
                classe_facturee = str(row.get('ClasseFacturee', '')).strip()
                classe_sortie   = str(row.get('ClasseDetectee_Sortie', '')).strip()
                date_sortie     = str(row.get('DateHeureSortie', '')).strip()

                img_binary = await images_map[img_key].read()

                try:
                    res = await analyze_transaction(
                        tx_id, paye, img_binary, g_entree, g_sortie,
                        classe_facturee, classe_sortie, date_sortie
                    )
                except QuotaEpuiseeError as qe:
                    yield f"data: {json.dumps({'type': 'error', 'message': f'Quota API épuisé après {idx} images. Rechargez vos crédits OpenAI.'})}\n\n"
                    return

                progress = int((idx + 1) / total_with * 100) if total_with else 100
                event = json.dumps({"type": "result", "progress": progress, "data": res}, allow_nan=False, default=str)
                yield f"data: {event}\n\n"

            # --- 5. Lignes sans image (dernier chunk uniquement) ---
            if is_last_chunk and rows_without_img:
                no_img_data = []
                for row in rows_without_img:
                    tx_id = str(row.get('NumeroSortie', row.get('NumeroTransaction', ''))).strip()
                    paye_val = row.get('TotalPaye', row.get('Espece', 0))
                    try:
                        paye = int(float(str(paye_val).replace(',', '.')))
                    except (ValueError, TypeError):
                        paye = 0
                    no_img_data.append({
                        "id": tx_id,
                        "classe_agent": f"{paye} FCFA",
                        "classe_facturee": str(row.get('ClasseFacturee', 'N/A')).strip(),
                        "gare_entree": str(row.get('NomGareEntree', '')).strip(),
                        "gare_sortie": str(row.get('NomGareSortie', '')).strip(),
                        "date_sortie": str(row.get('DateHeureSortie', '')).strip(),
                        "photo_attendue": str(row.get('PhotoSortie', '')).strip(),
                    })
                event = json.dumps({"type": "no_image_rows", "data": no_img_data}, allow_nan=False, default=str)
                yield f"data: {event}\n\n"

            print(f"[AUDIT] Chunk {chunk_index} terminé en {time.time()-t_start:.1f}s", flush=True)
            yield f"data: {json.dumps({'type': 'done', 'total_with': total_with, 'total_without': total_without})}\n\n"

        except Exception as e:
            print(f"[CRASH] Chunk {chunk_index} | {type(e).__name__}: {e}", flush=True)
            traceback.print_exc(file=sys.stdout)
            sys.stdout.flush()
            yield f"data: {json.dumps({'type': 'error', 'message': f'{type(e).__name__}: {e}'})}\n\n"
        finally:
            await form_data.close()

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


# ─────────────────────────────────────────────
#  ANALYSE RECETTES — endpoints
# ─────────────────────────────────────────────

BASE_DIRS = {
    "AMT": r"D:\AMT",
    "TT":  r"D:\TT",
}


def _station_name(dir_name: str) -> str:
    parts = dir_name.split('-', 1)
    return parts[1].strip() if len(parts) > 1 else dir_name


def _load_station_csvs(station_dir: str, years: list[int]) -> pd.DataFrame:
    files = sorted(glob_module.glob(os.path.join(station_dir, "*.csv")))
    dfs = []
    for f in files:
        m = re.search(r'(\d{4})', os.path.basename(f))
        year = int(m.group(1)) if m else None
        if years and year not in years:
            continue
        try:
            df = pd.read_csv(f, sep=None, engine='python')
            df.columns = [c.strip() for c in df.columns]
            df['_year'] = year
            dfs.append(df)
        except Exception as e:
            print(f"[CSV] Erreur lecture {f}: {e}", flush=True)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def _aggregate(df: pd.DataFrame, granularite: str) -> tuple[list, dict]:
    for col in ['Espece', 'ETCPaye', 'TotalPaye']:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(',', '.', regex=False),
                errors='coerce'
            ).fillna(0)

    df['_date'] = pd.to_datetime(df['DateHeureSortie'], dayfirst=True, errors='coerce')
    df = df.dropna(subset=['_date'])

    fmt = {'jour': '%Y-%m-%d', 'mois': '%Y-%m', 'annee': '%Y'}
    df['Periode'] = df['_date'].dt.strftime(fmt.get(granularite, '%Y-%m'))

    rows = []
    for periode, g in df.groupby('Periode'):
        row = {
            'periode': periode,
            'nb_transactions': len(g),
            'recette_espece':  int(g['Espece'].sum())   if 'Espece'   in g.columns else 0,
            'recette_etc':     int(g['ETCPaye'].sum())  if 'ETCPaye'  in g.columns else 0,
            'recette_totale':  int(g['TotalPaye'].sum())if 'TotalPaye' in g.columns else 0,
        }
        for cl in ['C1', 'C2', 'C3', 'C4']:
            mask = g.get('ClasseFacturee', pd.Series(dtype=str)) == cl
            row[f'nb_{cl}']      = int(mask.sum())
            row[f'recette_{cl}'] = int(g.loc[mask, 'TotalPaye'].sum()) if 'TotalPaye' in g.columns else 0
        rows.append(row)

    rows.sort(key=lambda x: x['periode'])

    totals = {k: sum(r.get(k, 0) for r in rows)
              for k in ['nb_transactions', 'recette_espece', 'recette_etc', 'recette_totale',
                        'nb_C1', 'nb_C2', 'nb_C3', 'nb_C4',
                        'recette_C1', 'recette_C2', 'recette_C3', 'recette_C4']}
    return rows, totals


@app.get("/stations")
async def get_stations():
    stations = []
    for reseau, base_dir in BASE_DIRS.items():
        if not os.path.isdir(base_dir):
            continue
        for entry in sorted(os.listdir(base_dir)):
            path = os.path.join(base_dir, entry)
            if not os.path.isdir(path):
                continue
            files = glob_module.glob(os.path.join(path, "*.csv"))
            years = sorted({int(m.group(1))
                            for f in files
                            if (m := re.search(r'(\d{4})', os.path.basename(f)))})
            stations.append({
                "id":     f"{reseau}_{entry}",
                "name":   _station_name(entry),
                "reseau": reseau,
                "dir":    path,
                "years":  years,
            })
    return {"stations": stations}


@app.post("/recettes")
async def compute_recettes(request: Request):
    body        = await request.json()
    station_dir = body.get("dir", "")
    years       = [int(y) for y in body.get("years", [])]
    granularite = body.get("granularite", "mois")

    if not os.path.isdir(station_dir):
        return JSONResponse({"error": "Répertoire introuvable"}, status_code=400)

    df = _load_station_csvs(station_dir, years)
    if df.empty:
        return JSONResponse({"error": "Aucun fichier CSV chargé"}, status_code=400)

    rows, totals = _aggregate(df, granularite)
    return {"data": rows, "totals": totals}


@app.post("/comparer")
async def comparer_rapport(request: Request):
    form        = await request.form(max_files=10, max_fields=20)
    rapport     = form.get("rapport")
    station_dir = form.get("dir", "")
    years       = json.loads(form.get("years", "[]"))
    granularite = form.get("granularite", "annee")

    if not os.path.isdir(station_dir):
        return JSONResponse({"error": "Répertoire introuvable"}, status_code=400)

    df_csv = _load_station_csvs(station_dir, [int(y) for y in years])
    if df_csv.empty:
        return JSONResponse({"error": "Aucun CSV chargé"}, status_code=400)

    _, totals = _aggregate(df_csv, granularite)

    content = await rapport.read()
    try:
        df_r = pd.read_excel(io.BytesIO(content), engine='openpyxl')
        df_r.columns = [str(c).strip() for c in df_r.columns]
    except Exception as e:
        return JSONResponse({"error": f"Erreur lecture Excel : {e}"}, status_code=400)

    def _find_col(df, *keywords):
        for kw in keywords:
            for col in df.columns:
                if kw.lower() in col.lower():
                    return col
        return None

    def _val(df, *kw):
        col = _find_col(df, *kw)
        if col is None:
            return None
        try:
            return float(pd.to_numeric(df[col], errors='coerce').sum())
        except Exception:
            return None

    r_vehicules = _val(df_r, 'vehicule', 'total', 'trafic', 'nb')
    r_etc       = _val(df_r, 'etc')
    r_mtc       = _val(df_r, 'mtc', 'espece', 'cash', 'manuel')
    r_recette   = _val(df_r, 'recette', 'montant', 'chiffre')
    r_c1        = _val(df_r, 'c1')
    r_c2        = _val(df_r, 'c2')
    r_c3        = _val(df_r, 'c3')
    r_c4        = _val(df_r, 'c4')

    def ecart(csv_val, rap_val):
        if rap_val is None:
            return None
        return round(csv_val - rap_val, 2)

    nb_total = totals['nb_transactions']
    nb_etc   = totals['recette_etc']   # nb payé ETC ≈ ETCPaye count
    nb_mtc   = totals['recette_espece']

    tableau = [
        {
            "indicateur": "Nombre de véhicules",
            "csv":        nb_total,
            "rapport":    r_vehicules,
            "ecart":      ecart(nb_total, r_vehicules),
        },
        {
            "indicateur": "Véhicules ETC",
            "csv":        totals.get('recette_etc', 0),
            "rapport":    r_etc,
            "ecart":      ecart(totals.get('recette_etc', 0), r_etc),
        },
        {
            "indicateur": "Véhicules MTC (Espèces)",
            "csv":        totals.get('recette_espece', 0),
            "rapport":    r_mtc,
            "ecart":      ecart(totals.get('recette_espece', 0), r_mtc),
        },
        {
            "indicateur": "Recette brute totale (FCFA)",
            "csv":        totals['recette_totale'],
            "rapport":    r_recette,
            "ecart":      ecart(totals['recette_totale'], r_recette),
        },
        {
            "indicateur": "Véhicules C1",
            "csv":        totals['nb_C1'],
            "rapport":    r_c1,
            "ecart":      ecart(totals['nb_C1'], r_c1),
        },
        {
            "indicateur": "Véhicules C2",
            "csv":        totals['nb_C2'],
            "rapport":    r_c2,
            "ecart":      ecart(totals['nb_C2'], r_c2),
        },
        {
            "indicateur": "Véhicules C3",
            "csv":        totals['nb_C3'],
            "rapport":    r_c3,
            "ecart":      ecart(totals['nb_C3'], r_c3),
        },
        {
            "indicateur": "Véhicules C4",
            "csv":        totals['nb_C4'],
            "rapport":    r_c4,
            "ecart":      ecart(totals['nb_C4'], r_c4),
        },
    ]

    nb_safe = nb_total if nb_total else 1
    mix_csv = {cl: round(totals[f'nb_{cl}'] / nb_safe * 100, 1) for cl in ['C1','C2','C3','C4']}
    mix_r   = {}
    if r_c1 is not None and r_c2 is not None and r_c3 is not None and r_c4 is not None:
        total_r = (r_c1 + r_c2 + r_c3 + r_c4) or 1
        mix_r = {cl: round(v / total_r * 100, 1) for cl, v in zip(['C1','C2','C3','C4'], [r_c1,r_c2,r_c3,r_c4])}
        for cl in ['C1','C2','C3','C4']:
            tableau.append({
                "indicateur": f"Mix {cl} (%)",
                "csv":        mix_csv[cl],
                "rapport":    mix_r.get(cl),
                "ecart":      ecart(mix_csv[cl], mix_r.get(cl)),
            })

    return {"tableau": tableau, "colonnes_rapport": list(df_r.columns)}


# ─────────────────────────────────────────────
#  ANALYSE EXCEL — lecture streaming, zéro crash
# ─────────────────────────────────────────────

def _safe_num(v) -> float:
    try:
        return float(str(v).replace(',', '.').replace(' ', '').strip())
    except Exception:
        return 0.0


def _parse_periode(val, granularite: str) -> str | None:
    """Parse a date cell and return the period string. Fast-path for DD/MM/YYYY."""
    s = str(val).strip() if val is not None else ''
    if not s or s in ('nan', 'None', 'NaT', ''):
        return None
    try:
        # Fast path: DD/MM/YYYY HH:MM:SS  or  DD/MM/YYYY
        if '/' in s:
            p = s.split('/')
            if len(p) >= 3:
                d, m, rest = p[0].zfill(2), p[1].zfill(2), p[2].split()[0]
                y = rest if len(rest) == 4 else rest[:4]
                if granularite == 'jour':  return f"{y}-{m}-{d}"
                if granularite == 'mois':  return f"{y}-{m}"
                return y
            # Variante fréquente: MM/YYYY ou M/YYYY
            if len(p) == 2 and p[0].strip().isdigit() and p[1].strip()[:4].isdigit():
                m = p[0].strip().zfill(2)
                y = p[1].strip()[:4]
                if granularite == 'jour':  return f"{y}-{m}-01"
                if granularite == 'mois':  return f"{y}-{m}"
                return y
    except Exception:
        pass
    # Excel serial date (nombre de jours depuis 1899-12-30)
    try:
        if isinstance(val, (int, float)) and val > 10_000:
            dt = pd.to_datetime(val, unit='D', origin='1899-12-30', errors='coerce')
            if pd.notna(dt):
                fmt = {'jour': '%Y-%m-%d', 'mois': '%Y-%m', 'annee': '%Y'}
                return dt.strftime(fmt.get(granularite, '%Y-%m'))
    except Exception:
        pass
    # Fallback: pandas
    try:
        dt = pd.to_datetime(s, dayfirst=True)
        fmt = {'jour': '%Y-%m-%d', 'mois': '%Y-%m', 'annee': '%Y'}
        return dt.strftime(fmt.get(granularite, '%Y-%m'))
    except Exception:
        return None


class Aggregateur:
    """Agrège les données à la volée, sans stocker toutes les lignes."""

    DATE_COLS   = ('DateHeureSortie', 'DateSortie', 'Date_Sortie', 'DateTransaction',
                   'Date', 'Heure_Sortie', 'DateHeure', 'Datetime', 'Timestamp',
                   # Variantes fréquentes (exports Excel / rapports)
                   'Date Heure Sortie', 'Date/Heure Sortie', 'Date Heure', 'Date/Heure',
                   'Date Sortie', 'Date de sortie', 'Date_Heure_Sortie')
    ESPECE_COLS = ('Espece', 'Especes', 'MontantEspece', 'Montant_Espece', 'RecetteEspece',
                   'Recette_Espece', 'PayeEspece', 'Paye_Espece', 'Espece_Paye',
                   'Cash', 'Liquide', 'MTC', 'MontantMTC', 'Montant_MTC',
                   'RecetteMTC', 'Recette_MTC', 'ManuelPaye', 'Manuel', 'Manuelle',
                   'EspecePaye', 'MTCPaye', 'MTC_Paye', 'CashPaye')
    ETC_COLS    = ('ETCPaye', 'ETC_Paye', 'MontantETC', 'Montant_ETC', 'RecetteETC',
                   'Recette_ETC', 'PayeETC', 'Paye_ETC', 'ETC_Montant', 'ETCMontant',
                   'ETC', 'Electronique', 'Badge', 'Electronic', 'BadgePaye',
                   'TelepeagePaye', 'Telepeage', 'Paiement_ETC', 'PaiementETC')
    TOTAL_COLS  = ('TotalPaye', 'Total_Paye', 'MontantTotal', 'Montant_Total',
                   'TotalMontant', 'Total', 'Montant', 'RecetteTotale', 'Recette_Totale',
                   'Recette', 'Amount', 'AmountTotal', 'TotalAmount',
                   # Variantes fréquentes
                   'Total Paye', 'Total Payé', 'Montant Total', 'Montant total', 'Recette totale')
    CLASSE_COLS = ('ClasseFacturee', 'Classe_Facturee', 'ClasseDetectee', 'ClasseDetectee_Sortie',
                   'Classe', 'Categorie', 'Category', 'TypeVehicule', 'Type_Vehicule',
                   'CategorieVehicule', 'Categorie_Vehicule', 'ClasseVehicule')
    MODE_COLS   = ('ModePaiement', 'Mode_Paiement', 'TypePaiement', 'Type_Paiement',
                   'ModeReglement', 'Mode_Reglement', 'TypeTransaction', 'PaymentMode',
                   'ModePayment', 'TypePayment')

    def __init__(self, granularite: str):
        self.gran   = granularite
        self.groups: dict = {}   # periode -> counters dict
        self._date_col   = None
        self._esp_col    = None
        self._etc_col    = None
        self._tot_col    = None
        self._cl_col     = None
        self._mode_col   = None  # colonne mode de paiement (fallback)
        self.detected_cols: dict = {}  # pour diagnostic

    @staticmethod
    def _norm_header(s: str) -> str:
        """
        Normalise un en-tête pour matcher malgré accents / espaces / ponctuation.
        Ex: "Total Payé" -> "totalpaye", "Date/Heure Sortie" -> "dateheuresortie"
        """
        if s is None:
            return ""
        s = str(s).strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        # garder uniquement alphanum (supprime espaces, _, -, /, etc.)
        s = re.sub(r"[^a-z0-9]+", "", s)
        return s

    def _detect_cols(self, headers: list[str]):
        # Map normalisée -> header original (premier gagnant)
        norm_map: dict[str, str] = {}
        for h in headers:
            h_orig = str(h).strip()
            key = self._norm_header(h_orig)
            if key and key not in norm_map:
                norm_map[key] = h_orig

        def pick(*names):
            # Match normalisé (exact puis substring) pour tolérer les variations.
            needles = [self._norm_header(n) for n in names if n]
            if not needles:
                return None
            # Exact d'abord
            for nd in needles:
                if nd in norm_map:
                    return norm_map[nd]
            # Substring ensuite (needle dans header ou header dans needle)
            for nd in needles:
                for hk, h_orig in norm_map.items():
                    if nd and (nd in hk or hk in nd):
                        return h_orig
            return None

        self._date_col = pick(*self.DATE_COLS)
        self._esp_col  = pick(*self.ESPECE_COLS)
        self._etc_col  = pick(*self.ETC_COLS)
        self._tot_col  = pick(*self.TOTAL_COLS)
        self._cl_col   = pick(*self.CLASSE_COLS)
        self._mode_col = pick(*self.MODE_COLS)
        self.detected_cols = {
            'date':   self._date_col,
            'espece': self._esp_col,
            'etc':    self._etc_col,
            'total':  self._tot_col,
            'classe': self._cl_col,
            'mode':   self._mode_col,
            'all_headers': list(headers),
        }
        print(f"[AGG] Colonnes détectées: date={self._date_col}, "
              f"esp={self._esp_col}, etc={self._etc_col}, "
              f"tot={self._tot_col}, cl={self._cl_col}, mode={self._mode_col}", flush=True)

    def _split_by_mode(self, mode_val: str, tot: float) -> tuple[float, float]:
        """Retourne (esp, etc) selon le mode de paiement."""
        m = str(mode_val).strip().upper()
        etc_keywords = ('ETC', 'ELECT', 'BADGE', 'TELEPEAGE', 'TÉLÉ', 'TELE', 'ELECTRONIC', 'TP')
        esp_keywords = ('ESPECE', 'ESPÈCE', 'MTC', 'CASH', 'MANUEL', 'LIQUID', 'MONNAIE')
        if any(k in m for k in etc_keywords):
            return 0.0, tot
        if any(k in m for k in esp_keywords):
            return tot, 0.0
        return 0.0, 0.0

    def add_row(self, row: dict):
        if self._date_col is None:
            self._detect_cols(list(row.keys()))
        periode = _parse_periode(row.get(self._date_col), self.gran)
        if periode is None:
            return
        if periode not in self.groups:
            self.groups[periode] = dict(nb=0, esp=0.0, etc=0.0, tot=0.0,
                                        nb_C1=0, nb_C2=0, nb_C3=0, nb_C4=0,
                                        nb_autres=0,
                                        r_C1=0.0, r_C2=0.0, r_C3=0.0, r_C4=0.0,
                                        r_autres=0.0)
        g = self.groups[periode]
        g['nb']  += 1
        esp = _safe_num(row.get(self._esp_col, 0)) if self._esp_col else 0
        etc = _safe_num(row.get(self._etc_col, 0)) if self._etc_col else 0
        tot = _safe_num(row.get(self._tot_col, 0)) if self._tot_col else 0
        # Fallback mode de paiement si esp+etc == 0 mais tot > 0
        if esp == 0 and etc == 0 and tot > 0 and self._mode_col:
            esp, etc = self._split_by_mode(row.get(self._mode_col, ''), tot)
        g['esp'] += esp
        g['etc'] += etc
        g['tot'] += tot
        cl = str(row.get(self._cl_col, '')).strip() if self._cl_col else ''
        if cl in ('C1', 'C2', 'C3', 'C4'):
            g[f'nb_{cl}'] += 1
            g[f'r_{cl}']  += tot
        else:
            # Toute catégorie inconnue (C?0, vide, etc.) → Autres
            g['nb_autres'] += 1
            g['r_autres']  += tot

    def add_chunk(self, df: pd.DataFrame):
        """Fast-path: ajoute un DataFrame entier (pour CSV chunked)."""
        if self._date_col is None:
            self._detect_cols(list(df.columns))
        date_col  = self._date_col  or ''
        esp_col   = self._esp_col   or ''
        etc_col   = self._etc_col   or ''
        tot_col   = self._tot_col   or ''
        cl_col    = self._cl_col    or ''
        mode_col  = self._mode_col  or ''
        if date_col not in df.columns:
            return
        fmt = {'jour': '%Y-%m-%d', 'mois': '%Y-%m', 'annee': '%Y'}.get(self.gran, '%Y-%m')
        df = df.copy()
        # Parsing dates: d'abord pandas (rapide), puis fallback ligne-à-ligne pour les formats exotiques
        raw_dates = df[date_col]
        dt = pd.to_datetime(raw_dates.astype(str), dayfirst=True, errors='coerce')
        if dt.isna().any():
            # Fallback uniquement sur les lignes non parsées
            mask = dt.isna()
            fallback = raw_dates[mask].apply(lambda v: _parse_periode(v, self.gran))
            # Reconstruire une colonne période hybride
            per = pd.Series(index=df.index, dtype="object")
            per[~mask] = dt[~mask].dt.strftime(fmt)
            per[mask] = fallback
            df['_per'] = per
            df = df.dropna(subset=['_per'])
        else:
            df['_dt'] = dt
            df['_per'] = df['_dt'].dt.strftime(fmt)
            df = df.dropna(subset=['_per'])

        for col in [esp_col, etc_col, tot_col]:
            if col and col in df.columns:
                df[col] = pd.to_numeric(
                    df[col].astype(str).str.replace(',', '.', regex=False), errors='coerce').fillna(0)
        for per, g in df.groupby('_per'):
            if per not in self.groups:
                self.groups[per] = dict(nb=0, esp=0.0, etc=0.0, tot=0.0,
                                        nb_C1=0, nb_C2=0, nb_C3=0, nb_C4=0,
                                        nb_autres=0,
                                        r_C1=0.0, r_C2=0.0, r_C3=0.0, r_C4=0.0,
                                        r_autres=0.0)
            d = self.groups[per]
            d['nb']  += len(g)
            tot_sum = float(g[tot_col].sum()) if tot_col in g.columns else 0
            esp_sum = float(g[esp_col].sum()) if esp_col in g.columns else 0
            etc_sum = float(g[etc_col].sum()) if etc_col in g.columns else 0

            # Fallback : si on a une colonne ModePaiement et que esp/etc sont vides
            if esp_sum == 0 and etc_sum == 0 and tot_sum > 0 and mode_col in g.columns:
                etc_kw = ('ETC', 'ELECT', 'BADGE', 'TELEPEAGE', 'TELE', 'TÉLÉ', 'TP')
                esp_kw = ('ESPECE', 'ESPÈCE', 'MTC', 'CASH', 'MANUEL', 'LIQUID')
                modes = g[mode_col].astype(str).str.upper()
                mask_etc = modes.apply(lambda m: any(k in m for k in etc_kw))
                mask_esp = modes.apply(lambda m: any(k in m for k in esp_kw))
                if tot_col in g.columns:
                    etc_sum = float(g.loc[mask_etc, tot_col].sum())
                    esp_sum = float(g.loc[mask_esp, tot_col].sum())

            d['esp'] += esp_sum
            d['etc'] += etc_sum
            d['tot'] += tot_sum
            if cl_col in g.columns:
                mask_known = g[cl_col].isin(['C1', 'C2', 'C3', 'C4'])
                for cl in ('C1', 'C2', 'C3', 'C4'):
                    mask = g[cl_col] == cl
                    d[f'nb_{cl}'] += int(mask.sum())
                    d[f'r_{cl}']  += float(g.loc[mask, tot_col].sum()) if tot_col in g.columns else 0
                # Catégories inconnues (C?0, vide, etc.) → Autres
                mask_autres = ~mask_known
                d['nb_autres'] += int(mask_autres.sum())
                d['r_autres']  += float(g.loc[mask_autres, tot_col].sum()) if tot_col in g.columns else 0
            else:
                # Pas de colonne classe : toutes les lignes sont "autres"
                d['nb_autres'] += len(g)

    def result(self) -> tuple[list, dict]:
        rows = [
            {
                'periode': per,
                'nb_transactions': g['nb'],
                'recette_espece':  int(g['esp']),
                'recette_etc':     int(g['etc']),
                'recette_totale':  int(g['tot']),
                'nb_C1': g['nb_C1'], 'nb_C2': g['nb_C2'],
                'nb_C3': g['nb_C3'], 'nb_C4': g['nb_C4'],
                'nb_autres': g['nb_autres'],
                'recette_C1': int(g['r_C1']), 'recette_C2': int(g['r_C2']),
                'recette_C3': int(g['r_C3']), 'recette_C4': int(g['r_C4']),
                'recette_autres': int(g['r_autres']),
            }
            for per, g in sorted(self.groups.items())
        ]
        keys = ['nb_transactions', 'recette_espece', 'recette_etc', 'recette_totale',
                'nb_C1', 'nb_C2', 'nb_C3', 'nb_C4', 'nb_autres',
                'recette_C1', 'recette_C2', 'recette_C3', 'recette_C4', 'recette_autres']
        totals = {k: sum(r.get(k, 0) for r in rows) for k in keys}
        return rows, totals


def _cache_dir() -> Path:
    # Par défaut: backend/.cache (persistant sur disque)
    return Path(os.getenv("CACHE_DIR", Path(__file__).parent / ".cache"))


def _cache_key(content: bytes, granularite: str) -> str:
    h = hashlib.sha256()
    h.update(b"analyser-excel|v1|")
    h.update(granularite.encode("utf-8", errors="ignore"))
    h.update(b"|")
    h.update(content)
    return h.hexdigest()


def _cache_get(key: str) -> dict | None:
    p = _cache_dir() / f"{key}.json"
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return None


def _cache_set(key: str, payload: dict) -> None:
    d = _cache_dir()
    d.mkdir(parents=True, exist_ok=True)
    p = d / f"{key}.json"
    tmp = d / f"{key}.tmp"
    tmp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    tmp.replace(p)


@app.post("/analyser-excel")
async def analyser_excel(
    fichier: UploadFile = File(...),
    granularite: str    = Form("mois"),
):
    """Traite UN fichier (xlsx/xls/csv) et retourne ses agrégats via SSE."""
    from openpyxl import load_workbook

    async def generate():
        nom = fichier.filename
        ext = os.path.splitext(nom)[1].lower()
        agg = Aggregateur(granularite)
        n_rows: int = 0
        n_errors: int = 0

        yield f"data: {json.dumps({'type': 'fichier_debut', 'nom': nom})}\n\n"

        try:
            content = await fichier.read()
            # Cache persistant: si le fichier (contenu) et la granularité n'ont pas changé,
            # on renvoie immédiatement les résultats.
            key = _cache_key(content, granularite)
            cached = _cache_get(key)
            if cached is not None:
                yield f"data: {json.dumps({'type': 'cache_hit', 'nom': nom})}\n\n"
                yield f"data: {json.dumps({'type': 'fichier_done', 'nom': nom, 'rows': cached.get('rows', 0), 'errors': cached.get('errors', 0)})}\n\n"
                if cached.get("detected"):
                    yield f"data: {json.dumps({'type': 'colonnes', 'detected': cached.get('detected')})}\n\n"
                yield f"data: {json.dumps({'type': 'resultat', 'data': cached.get('data', []), 'totals': cached.get('totals', {})})}\n\n"
                yield f"data: {json.dumps({'type': 'done'})}\n\n"
                return

            if ext in ('.xlsx', '.xlsm', '.xlam'):
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                headers = None
                for xrow in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c).strip() if c is not None else f'_col{i}'
                                   for i, c in enumerate(xrow)]
                        agg._detect_cols(headers)
                        continue
                    try:
                        agg.add_row(dict(zip(headers, xrow)))
                        n_rows += 1
                    except Exception:
                        n_errors += 1
                    if n_rows % 5_000 == 0 and n_rows:
                        yield f"data: {json.dumps({'type': 'progress', 'nom': nom, 'rows': n_rows})}\n\n"
                wb.close()

            elif ext == '.xls':
                df = pd.read_excel(io.BytesIO(content), engine='xlrd')
                df.columns = [str(c).strip() for c in df.columns]
                agg.add_chunk(df)
                n_rows = len(df)

            elif ext == '.csv':
                for chunk in pd.read_csv(io.BytesIO(content), sep=None,
                                         engine='python', chunksize=10_000):
                    chunk.columns = [str(c).strip() for c in chunk.columns]
                    agg.add_chunk(chunk)
                    n_rows += len(chunk)
                    yield f"data: {json.dumps({'type': 'progress', 'nom': nom, 'rows': n_rows})}\n\n"

            else:
                yield f"data: {json.dumps({'type': 'erreur', 'nom': nom, 'message': 'Format non supporté (.xlsx .xls .csv)'})}\n\n"
                yield f"data: {json.dumps({'type': 'done'})}\n\n"
                return

            rows, totals = agg.result()
            yield f"data: {json.dumps({'type': 'fichier_done', 'nom': nom, 'rows': n_rows, 'errors': n_errors})}\n\n"
            yield f"data: {json.dumps({'type': 'colonnes', 'detected': agg.detected_cols})}\n\n"
            yield f"data: {json.dumps({'type': 'resultat', 'data': rows, 'totals': totals})}\n\n"
            _cache_set(key, {
                "key": key,
                "nom": nom,
                "granularite": granularite,
                "rows": n_rows,
                "errors": n_errors,
                "detected": agg.detected_cols,
                "data": rows,
                "totals": totals,
                "cached_at": time.time(),
            })

        except Exception as e:
            traceback.print_exc()
            yield f"data: {json.dumps({'type': 'erreur', 'nom': nom, 'message': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/comparer-totaux")
async def comparer_totaux(request: Request):
    """Compare des totaux déjà calculés avec un fichier Excel rapport d'exploitation."""
    form    = await request.form(max_files=5, max_fields=20)
    rapport = form.get("rapport")
    totals  = json.loads(form.get("totals", "{}"))

    content = await rapport.read()
    try:
        df_r = pd.read_excel(io.BytesIO(content), engine='openpyxl')
        df_r.columns = [str(c).strip() for c in df_r.columns]
    except Exception as e:
        return JSONResponse({"error": f"Erreur lecture Excel : {e}"}, status_code=400)

    def _find(df, *kws):
        for kw in kws:
            for col in df.columns:
                if kw.lower() in col.lower():
                    return col
        return None

    def _val(df, *kws):
        col = _find(df, *kws)
        if col is None:
            return None
        try:
            return float(pd.to_numeric(df[col], errors='coerce').sum())
        except Exception:
            return None

    def ec(csv_v, r_v):
        return round(float(csv_v) - r_v, 2) if r_v is not None else None

    r_veh  = _val(df_r, 'vehicule', 'trafic', 'total_veh', 'nb_veh')
    r_etc  = _val(df_r, 'etc', 'electronique', 'badge')
    r_mtc  = _val(df_r, 'mtc', 'espece', 'cash', 'manuel', 'liquide')
    r_rec  = _val(df_r, 'recette', 'montant', 'chiffre', 'revenue', 'total_rec')
    r_c    = {cl: _val(df_r, cl) for cl in ('C1', 'C2', 'C3', 'C4')}

    nb  = totals.get('nb_transactions', 0)
    tableau = [
        {"indicateur": "Nombre de véhicules",       "csv": nb,                              "rapport": r_veh,  "ecart": ec(nb,                              r_veh)},
        {"indicateur": "Véhicules ETC",             "csv": totals.get('recette_etc', 0),    "rapport": r_etc,  "ecart": ec(totals.get('recette_etc', 0),    r_etc)},
        {"indicateur": "Véhicules MTC (Espèces)",   "csv": totals.get('recette_espece', 0), "rapport": r_mtc,  "ecart": ec(totals.get('recette_espece', 0), r_mtc)},
        {"indicateur": "Recette brute totale (FCFA)","csv": totals.get('recette_totale', 0), "rapport": r_rec,  "ecart": ec(totals.get('recette_totale', 0), r_rec)},
        {"indicateur": "Véhicules C1",              "csv": totals.get('nb_C1', 0),          "rapport": r_c['C1'], "ecart": ec(totals.get('nb_C1', 0),       r_c['C1'])},
        {"indicateur": "Véhicules C2",              "csv": totals.get('nb_C2', 0),          "rapport": r_c['C2'], "ecart": ec(totals.get('nb_C2', 0),       r_c['C2'])},
        {"indicateur": "Véhicules C3",              "csv": totals.get('nb_C3', 0),          "rapport": r_c['C3'], "ecart": ec(totals.get('nb_C3', 0),       r_c['C3'])},
        {"indicateur": "Véhicules C4",              "csv": totals.get('nb_C4', 0),          "rapport": r_c['C4'], "ecart": ec(totals.get('nb_C4', 0),       r_c['C4'])},
    ]

    # Mix %
    nb_safe = nb or 1
    mix_csv = {cl: round(totals.get(f'nb_{cl}', 0) / nb_safe * 100, 1) for cl in ('C1','C2','C3','C4')}
    if all(r_c[cl] is not None for cl in ('C1','C2','C3','C4')):
        tot_r = sum(r_c[cl] for cl in ('C1','C2','C3','C4')) or 1
        for cl in ('C1','C2','C3','C4'):
            mix_r = round(r_c[cl] / tot_r * 100, 1)
            tableau.append({"indicateur": f"Mix {cl} (%)", "csv": mix_csv[cl],
                            "rapport": mix_r, "ecart": ec(mix_csv[cl], mix_r)})

    return {"tableau": tableau, "colonnes_rapport": list(df_r.columns)}


# ─────────────────────────────────────────────────────────────────
#  CONTRÔLE RECETTE POST PAID
# ─────────────────────────────────────────────────────────────────

# Colonnes plaque dans les données péage
PLAQUE_PEAGE_COLS = (
    'PlaqueLPR_Entree', 'PlaqueLPR_Sortie', 'PlaqueLPR',
    'Immatriculation', 'NumImmatriculation', 'NumeroImmatriculation',
    'ImmatriculationVehicule', 'Immat', 'Plaque', 'NumPlaque',
    'NumeroDePlaque', 'PlaqueMineralogique', 'VehicleRegistration',
    'Registration', 'ImmatVehicule', 'Matricule', 'NumMatricule',
    'NumeroVehicule', 'VehicleID',
)

# Colonnes plaque dans les factures post-paid
PLAQUE_FACT_COLS = PLAQUE_PEAGE_COLS + ('Vehicule', 'Vehicle', 'Numero', 'NumVehicule')

# Colonnes passages dans les factures
PASSAGES_FACT_COLS = (
    'NombrePassages', 'Nb_Passages', 'NbPassages', 'Passages', 'NombreVoyages',
    'Quantite', 'Qte', 'Qty', 'Nombre', 'NbTransactions', 'NbTx',
    'NombreTransactions', 'Transactions', 'Passage', 'Count',
)

# Colonnes montant dans les factures
MONTANT_FACT_COLS = (
    'Montant', 'MontantTTC', 'MontantHT', 'Total', 'TotalPaye', 'TotalTTC',
    'MontantTotal', 'Somme', 'MontantFacture', 'Recette', 'Prix',
    'Amount', 'AmountTotal', 'MontantDu', 'Net',
)

# Colonnes client/société dans les factures
CLIENT_FACT_COLS = (
    'NomClient', 'Client', 'Societe', 'Société', 'Entreprise',
    'RaisonSociale', 'Raison_Sociale', 'Nom', 'NomSociete', 'NomEntreprise',
    'ClientName', 'Company',
)

# Colonnes péage/gare dans les données péage
GARE_PEAGE_COLS = (
    'NomGareSortie', 'NomGareEntree', 'Gare', 'Station', 'GarePeage',
    'NomGare', 'LibelleGare', 'Peage',
)
GARE_ENTREE_PEAGE_COLS = ('NomGareEntree', 'GareEntree', 'Entree', 'StationEntree', 'LibelleGareEntree')
GARE_SORTIE_PEAGE_COLS = ('NomGareSortie', 'GareSortie', 'Sortie', 'StationSortie', 'LibelleGareSortie')
TARIF_PEAGE_COLS = ('TotalPaye', 'Total', 'Montant', 'TotalTTC', 'Tarif', 'Prix', 'MontantTotal', 'Recette', 'Espece')
DATE_PEAGE_COLS  = ('DateHeureSortie', 'DateHeureEntree', 'DateHeure', 'DateSortie', 'DateEntree', 'Date', 'DatePassage')
CLASSE_PEAGE_COLS = ('ClasseFacturee', 'ClasseDetectee_Entree', 'ClasseDetectee', 'Classe', 'Category', 'TypeVehicule', 'Categorie')

# Colonnes facture post-paid
MOIS_FACT_COLS   = ('Mois', 'Month', 'Periode', 'Mois_facturation', 'MoisFacturation')
PLAQUE_FACT2_COLS = ('Plaques', 'Plaque') + PLAQUE_PEAGE_COLS
CLIENT_FACT2_COLS = ('Nom du client', 'NomClient', 'Client', 'Societe', 'NomSociete', 'Entreprise', 'RaisonSociale')
MONTANT_FACT2_COLS = ('Montant total', 'MontantTotal', 'Montant', 'Total', 'TotalPaye', 'TotalTTC')

# Mapping mois français → numéro
_FR_MOIS: dict = {
    'jan': 1, 'janv': 1, 'janvier': 1,
    'fev': 2, 'fevr': 2, 'fevrier': 2,
    'mar': 3, 'mars': 3,
    'avr': 4, 'avril': 4,
    'mai': 5,
    'jun': 6, 'juin': 6,
    'jul': 7, 'juil': 7, 'juillet': 7,
    'aou': 8, 'aout': 8,
    'sep': 9, 'sept': 9, 'septembre': 9,
    'oct': 10, 'octobre': 10,
    'nov': 11, 'novembre': 11,
    'dec': 12, 'decembre': 12,
}


def _norm(s: str) -> str:
    """Normalise un header pour matching."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _normalize_plaque(p: str) -> str:
    """'AA 768 BC' → 'AA768BC', 'SL-6495-B' → 'SL6495B', 'A/B' → 'A'."""
    p = str(p).strip().upper()
    if '/' in p:
        p = p.split('/')[0].strip()
    return re.sub(r'[\s\-]+', '', p)


def _parse_mois_fr(mois_str: str) -> str:
    """'juil-21' → '2021-07', 'août-21' → '2021-08', '' → ''."""
    s = unicodedata.normalize('NFKD', mois_str.strip().lower())
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    parts = re.split(r'[-/\s]+', s)
    if len(parts) >= 2:
        month = _FR_MOIS.get(parts[0])
        if month:
            try:
                year = int(parts[1])
                if year < 100:
                    year += 2000
                return f"{year}-{month:02d}"
            except ValueError:
                pass
    m = re.match(r'^(\d{1,2})[/-](\d{4})$', s)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    m = re.match(r'^(\d{4})[/-](\d{1,2})$', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return ''


def _extract_mois_key(date_str: str) -> str:
    """'31/12/2021 23:59' → '2021-12',  '2021-12-31 ...' → '2021-12'."""
    if not date_str:
        return ''
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', date_str)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}"
    m = re.match(r'^(\d{4})-(\d{1,2})', date_str)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    return ''


def _pick_col(headers: list, *candidates) -> str | None:
    """Retourne le premier header qui matche un des candidats (exact puis substring)."""
    norm_map = {_norm(h): h for h in headers if h}
    needles = [_norm(c) for c in candidates if c]
    for nd in needles:
        if nd in norm_map:
            return norm_map[nd]
    for nd in needles:
        for hk, horig in norm_map.items():
            if nd and (nd in hk or hk in nd):
                return horig
    return None


def _safe_float(v) -> float:
    try:
        return float(str(v).replace(',', '.').replace(' ', '').replace('\xa0', ''))
    except Exception:
        return 0.0


def _safe_int(v) -> int:
    try:
        return int(round(_safe_float(v)))
    except Exception:
        return 0


def _read_file_to_df(content: bytes, filename: str) -> pd.DataFrame:
    """Lit un fichier xlsx/xls/csv et retourne un DataFrame."""
    ext = os.path.splitext(filename)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        return pd.read_excel(io.BytesIO(content), engine='openpyxl')
    elif ext == '.xls':
        return pd.read_excel(io.BytesIO(content), engine='xlrd')
    elif ext == '.csv':
        return pd.read_csv(io.BytesIO(content), sep=None, engine='python')
    else:
        raise ValueError(f"Format non supporté : {ext}")


def _find_header_row(df_raw: pd.DataFrame, plate_candidates: tuple) -> pd.DataFrame:
    """
    Cherche la ligne d'en-tête réelle dans un DataFrame brut (cas factures Excel avec
    des lignes de titre avant le tableau). Retourne un DataFrame avec les bons headers.
    """
    norm_cands = [_norm(c) for c in plate_candidates]
    for i, row in df_raw.iterrows():
        vals = [_norm(str(v)) for v in row.values if v is not None and str(v).strip()]
        if any(nd in v or v in nd for nd in norm_cands for v in vals if nd and v):
            # Cette ligne est probablement l'en-tête
            new_df = df_raw.iloc[i + 1:].copy()
            new_df.columns = [str(c).strip() if c is not None else f'_col{j}'
                              for j, c in enumerate(df_raw.iloc[i].values)]
            new_df = new_df.reset_index(drop=True)
            return new_df
    # Pas trouvé: retourner tel quel
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    return df_raw


def _parse_facture_file(content: bytes, filename: str) -> list[dict]:
    """
    Parse un fichier de facture post-paid.
    Retourne une liste de {client, plaque, passages_factures, montant_facture, gare}.
    """
    ext = os.path.splitext(filename)[1].lower()
    client_from_filename = os.path.splitext(os.path.basename(filename))[0]
    # Retirer le mois/année du nom de fichier pour obtenir le nom du client
    client_from_filename = re.sub(r'\s*(jan|fev|mar|avr|mai|juin|juil|aout|sep|oct|nov|dec|janvier|'
                                  r'fevrier|mars|avril|mai|juin|juillet|aout|septembre|octobre|'
                                  r'novembre|decembre)\.?\s*\d*', '',
                                  client_from_filename, flags=re.IGNORECASE).strip()
    client_from_filename = re.sub(r'\s*\d{4}\s*', '', client_from_filename).strip()
    client_from_filename = client_from_filename or os.path.splitext(os.path.basename(filename))[0]

    if ext in ('.xlsx', '.xlsm'):
        df_raw = pd.read_excel(io.BytesIO(content), engine='openpyxl', header=None)
    elif ext == '.xls':
        df_raw = pd.read_excel(io.BytesIO(content), engine='xlrd', header=None)
    elif ext == '.csv':
        df_raw = pd.read_csv(io.BytesIO(content), sep=None, engine='python', header=None)
    else:
        raise ValueError(f"Format non supporté : {ext}")

    df = _find_header_row(df_raw, PLAQUE_FACT_COLS + PASSAGES_FACT_COLS)

    col_plaque    = _pick_col(list(df.columns), *PLAQUE_FACT_COLS)
    col_passages  = _pick_col(list(df.columns), *PASSAGES_FACT_COLS)
    col_montant   = _pick_col(list(df.columns), *MONTANT_FACT_COLS)
    col_client    = _pick_col(list(df.columns), *CLIENT_FACT_COLS)
    col_gare      = _pick_col(list(df.columns), *GARE_PEAGE_COLS)

    rows = []
    for _, row in df.iterrows():
        plaque_val = str(row.get(col_plaque, '')).strip() if col_plaque else ''
        if not plaque_val or plaque_val.lower() in ('nan', 'none', '', 'total', 'sous-total'):
            continue
        # Ignorer les lignes de total/sous-total
        if re.match(r'^(total|sous[-\s]total|grand total|sous total)$', plaque_val, re.IGNORECASE):
            continue

        client_val   = str(row.get(col_client, '')).strip() if col_client else ''
        passages_val = _safe_int(row.get(col_passages, 1)) if col_passages else 1
        montant_val  = _safe_float(row.get(col_montant, 0)) if col_montant else 0.0
        gare_val     = str(row.get(col_gare, '')).strip() if col_gare else ''

        # Si pas de colonne client dédiée, utiliser le nom du fichier
        client_final = client_val if client_val and client_val.lower() not in ('nan', 'none') \
                       else client_from_filename

        rows.append({
            'client':            client_final,
            'plaque':            plaque_val.upper(),
            'passages_factures': passages_val,
            'montant_facture':   montant_val,
            'gare':              gare_val,
        })
    return rows


def _extract_passages_peage(content: bytes, filename: str) -> dict[str, list]:
    """
    Lit un fichier péage et retourne un dict {plaque_upper: [passage_dict, ...]}.
    Chaque passage_dict contient : gare_e, gare_s, date, montant, classe.
    """
    ext = os.path.splitext(filename)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        df = pd.read_excel(io.BytesIO(content), engine='openpyxl')
    elif ext == '.xls':
        df = pd.read_excel(io.BytesIO(content), engine='xlrd')
    elif ext == '.csv':
        df = pd.read_csv(io.BytesIO(content), sep=None, engine='python')
    else:
        raise ValueError(f"Format non supporté : {ext}")

    df.columns = [str(c).strip() for c in df.columns]
    cols = list(df.columns)
    col_plaque  = _pick_col(cols, *PLAQUE_PEAGE_COLS)
    col_gare_e  = _pick_col(cols, *GARE_ENTREE_PEAGE_COLS)
    col_gare_s  = _pick_col(cols, *GARE_SORTIE_PEAGE_COLS)
    col_tarif   = _pick_col(cols, *TARIF_PEAGE_COLS)
    col_date    = _pick_col(cols, *DATE_PEAGE_COLS)
    col_classe  = _pick_col(cols, *CLASSE_PEAGE_COLS)

    # Si pas de colonnes entrée/sortie séparées, fallback générique
    if not col_gare_e and not col_gare_s:
        col_gare_e = _pick_col(cols, *GARE_PEAGE_COLS)

    # Colonne secondaire pour les CSV à double champ LPR (entrée + sortie)
    col_plaque2 = None
    if col_plaque == 'PlaqueLPR_Entree' and 'PlaqueLPR_Sortie' in df.columns:
        col_plaque2 = 'PlaqueLPR_Sortie'
    elif col_plaque == 'PlaqueLPR_Sortie' and 'PlaqueLPR_Entree' in df.columns:
        col_plaque2 = 'PlaqueLPR_Entree'

    SKIP = {'NAN', 'NONE', '', 'NO PLAQUES', 'NOPLAQUES', 'NOPLAQUE', '-', 'N/A', 'NA'}

    passages: dict[str, list] = {}
    if col_plaque is None:
        return passages

    for _, row in df.iterrows():
        plaque = str(row.get(col_plaque, '')).strip().upper()
        if plaque in SKIP and col_plaque2:
            plaque = str(row.get(col_plaque2, '')).strip().upper()
        if not plaque or plaque in SKIP:
            continue

        entry: dict = {}
        if col_gare_e:
            entry['gare_e'] = str(row.get(col_gare_e, '')).strip()
        if col_gare_s:
            entry['gare_s'] = str(row.get(col_gare_s, '')).strip()
        if col_date:
            entry['date'] = str(row.get(col_date, '')).strip()
        if col_tarif:
            entry['montant'] = _safe_float(row.get(col_tarif))
        if col_classe:
            entry['classe'] = str(row.get(col_classe, '')).strip()

        if plaque not in passages:
            passages[plaque] = []
        passages[plaque].append(entry)

    return passages


def _pp_cache_key_peage(content: bytes) -> str:
    h = hashlib.sha256()
    h.update(b"postpaid-peage|v3|")  # v3: passage dicts with montant/date/classe
    h.update(content)
    return h.hexdigest()


def _read_plaques_file(content: bytes, filename: str) -> list[str]:
    """Lit un fichier liste de plaques et retourne la liste complète (avec doublons)."""
    df = _read_file_to_df(content, filename)
    df.columns = [str(c).strip() for c in df.columns]
    col = _pick_col(list(df.columns), *PLAQUE_PEAGE_COLS,
                    "Plaque d'immatriculation", 'Plaque d immatriculation')
    if col is None:
        col = df.columns[0]
    skip = {'nan', 'none', '', 'plaque', 'immatriculation', 'matricule',
            "plaque d'immatriculation", 'plaque d immatriculation'}
    plates = []
    for v in df[col]:
        p = str(v).strip().upper()
        if p and p.lower() not in skip:
            plates.append(p)
    return plates


@app.post("/controle-postpaid/facture")
async def read_facture_endpoint(request: Request):
    """
    Lit le fichier de facturation post-paid :
    colonnes Mois | Plaques | Nom du client | Montant total
    Retourne une liste de lignes {mois, mois_key, plaque, plaque_key, client, montant_facture}.
    """
    form = await request.form(max_files=1, max_fields=10)
    fichier = form.get("fichier_facture")
    if not fichier:
        return JSONResponse({"error": "Aucun fichier fourni"}, status_code=400)
    try:
        content = await fichier.read()
        df = _read_file_to_df(content, fichier.filename)
        df.columns = [str(c).strip() for c in df.columns]
        cols = list(df.columns)
        col_mois    = _pick_col(cols, *MOIS_FACT_COLS)
        col_plaque  = _pick_col(cols, *PLAQUE_FACT2_COLS)
        col_client  = _pick_col(cols, *CLIENT_FACT2_COLS)
        col_montant = _pick_col(cols, *MONTANT_FACT2_COLS)

        if col_plaque is None:
            col_plaque = cols[1] if len(cols) > 1 else cols[0]

        rows = []
        clients = set()
        for _, row in df.iterrows():
            mois_str = str(row.get(col_mois, '')).strip() if col_mois else ''
            plaque_raw = str(row.get(col_plaque, '')).strip()
            plaque_up  = plaque_raw.upper()
            client  = str(row.get(col_client, '')).strip()  if col_client  else ''
            montant = _safe_float(row.get(col_montant))     if col_montant else 0.0

            if not plaque_raw or plaque_up in ('NAN', 'NONE', ''):
                continue

            mois_key   = _parse_mois_fr(mois_str)
            plaque_key = _normalize_plaque(plaque_raw)
            rows.append({
                'mois':           mois_str,
                'mois_key':       mois_key,
                'plaque':         plaque_up,
                'plaque_key':     plaque_key,
                'client':         client,
                'montant_facture': montant,
            })
            if client:
                clients.add(client)

        return JSONResponse({
            'rows':    rows,
            'total':   len(rows),
            'clients': sorted(clients),
        })
    except Exception as e:
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/controle-postpaid/plaques")
async def read_plaques_endpoint(request: Request):
    """
    Lit un fichier liste de plaques (colonne unique) et retourne la liste avec comptages.
    Requête légère — fichier tiny, réponse JSON directe.
    """
    form = await request.form(max_files=1, max_fields=10)
    fichier = form.get("fichier_plaques")
    if not fichier:
        return JSONResponse({"error": "Aucun fichier fourni"}, status_code=400)
    try:
        content = await fichier.read()
        plaques_raw = _read_plaques_file(content, fichier.filename)
        counts: dict[str, int] = {}
        for p in plaques_raw:
            counts[p] = counts.get(p, 0) + 1
        result = [{"plaque": p, "nb_liste": n} for p, n in sorted(counts.items())]
        return JSONResponse({
            "plaques": result,
            "total_lignes": len(plaques_raw),
            "total_uniques": len(counts),
        })
    except Exception as e:
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/controle-postpaid/peage-file")
async def count_peage_file(request: Request):
    """
    Compte les passages par plaque dans UN seul fichier péage.
    Le body est lu AVANT de démarrer le StreamingResponse pour éviter
    les MultipartParseError liés au boundary.
    Cache SHA256 — deuxième upload du même fichier = réponse instantanée.
    """
    try:
        form = await request.form(max_files=1, max_fields=10)
        fichier = form.get("fichier_peage")
        if not fichier:
            return JSONResponse({"error": "Fichier manquant"}, status_code=400)
        nom = fichier.filename
        content = await fichier.read()
        # Plaques cibles : on ne renvoie le détail complet que pour ces plaques
        try:
            raw_cibles = form.get("plaques_cibles", "") or ""
            plaques_cibles: set = set(json.loads(raw_cibles)) if raw_cibles else set()
        except Exception:
            plaques_cibles = set()
    except Exception as e:
        traceback.print_exc()
        return JSONResponse({"error": f"Lecture du fichier impossible : {e}"}, status_code=400)

    async def generate():
        try:
            key = _pp_cache_key_peage(content)
            cached = _cache_get(key)

            if cached is not None:
                passages = cached.get('plaques', {})
                yield f"data: {json.dumps({'type': 'progress', 'message': f'⚡ {nom} (cache) : {len(passages)} plaques'})}\n\n"
            else:
                yield f"data: {json.dumps({'type': 'progress', 'message': f'Lecture {nom}…'})}\n\n"
                passages = _extract_passages_peage(content, nom)
                _cache_set(key, {'plaques': passages})
                yield f"data: {json.dumps({'type': 'progress', 'message': f'{nom} : {len(passages)} plaques trouvées'})}\n\n"

            counts: dict = {}
            details: dict = {}
            for plaque, plist in passages.items():
                counts[plaque] = len(plist)
                if plaque in plaques_cibles:
                    # Groupe par mois (YYYY-MM) en utilisant la date de sortie
                    by_mois: dict = {}
                    for p in plist:
                        if not isinstance(p, dict):
                            continue
                        mk = _extract_mois_key(p.get('date', ''))
                        if mk not in by_mois:
                            by_mois[mk] = {'montant': 0.0, 'nb': 0, 'passages': []}
                        by_mois[mk]['montant'] += p.get('montant', 0) or 0
                        by_mois[mk]['nb'] += 1
                        by_mois[mk]['passages'].append(p)
                    details[plaque] = by_mois

            yield f"data: {json.dumps({'type': 'resultat', 'counts': counts, 'details': details, 'fichier': nom})}\n\n"

        except Exception as e:
            traceback.print_exc()
            yield f"data: {json.dumps({'type': 'erreur', 'message': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── ancienne route (compatibilité) ──────────────────────────────────────────
@app.post("/controle-postpaid")
async def controle_postpaid(request: Request):
    """Redirige vers les nouveaux endpoints — conservé pour compatibilité."""
    async def generate():
        yield f"data: {json.dumps({'type': 'erreur', 'message': 'Utilisez /controle-postpaid/plaques et /controle-postpaid/peage-file'})}\n\n"
        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    async def generate():
        try:
            total_steps = len(fichiers_peage) + len(fichiers_factures) + 1  # +1 rapprochement
            step = 0

            def pct(extra=0):
                return min(99, int((step + extra) / total_steps * 95))

            def evt_pbar(p, msg=""):
                return f"data: {json.dumps({'type': 'progress_bar', 'pct': p, 'message': msg})}\n\n"

            yield evt_pbar(0, "Démarrage…")

            # ── 1. Lecture des données péage ──────────────────────────────
            yield f"data: {json.dumps({'type': 'info', 'message': f'Lecture des données péage ({len(fichiers_peage)} fichier(s))…'})}\n\n"

            passages_peage: dict[str, list] = {}
            peage_col_found = False

            for f in fichiers_peage:
                nom = f.filename
                try:
                    content = await f.read()
                    key = _pp_cache_key_peage(content)
                    cached = _cache_get(key)

                    if cached is not None:
                        partial = {p: v for p, v in cached.get('plaques', {}).items()}
                        yield f"data: {json.dumps({'type': 'progress', 'message': f'⚡ {nom} (cache) : {len(partial)} plaques'})}\n\n"
                    else:
                        partial = _extract_passages_peage(content, nom)
                        _cache_set(key, {'plaques': partial})
                        yield f"data: {json.dumps({'type': 'progress', 'message': f'{nom} : {len(partial)} plaques trouvées'})}\n\n"

                    if partial:
                        peage_col_found = True
                    for plaque, gares in partial.items():
                        if plaque not in passages_peage:
                            passages_peage[plaque] = []
                        passages_peage[plaque].extend(gares)

                except Exception as e:
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'Erreur lecture {nom}: {e}'})}\n\n"

                step += 1
                yield evt_pbar(pct(), nom)

            if not peage_col_found:
                yield f"data: {json.dumps({'type': 'warning', 'message': 'Colonne plaque non détectée dans les données péage.'})}\n\n"

            yield f"data: {json.dumps({'type': 'info', 'message': f'Péage : {len(passages_peage)} plaques distinctes chargées'})}\n\n"

            # ── 2. Lecture des factures post-paid ─────────────────────────
            yield f"data: {json.dumps({'type': 'info', 'message': f'Lecture des factures ({len(fichiers_factures)} fichier(s))…'})}\n\n"

            lignes_factures: list[dict] = []
            for f in fichiers_factures:
                nom = f.filename
                try:
                    content = await f.read()
                    key = _pp_cache_key_facture(content)
                    cached = _cache_get(key)

                    if cached is not None:
                        lignes = cached.get('rows', [])
                        yield f"data: {json.dumps({'type': 'progress', 'message': f'⚡ {nom} (cache) : {len(lignes)} lignes'})}\n\n"
                    else:
                        lignes = _parse_facture_file(content, nom)
                        _cache_set(key, {'rows': lignes})
                        yield f"data: {json.dumps({'type': 'progress', 'message': f'{nom} : {len(lignes)} lignes facturées'})}\n\n"

                    lignes_factures.extend(lignes)

                except Exception as e:
                    yield f"data: {json.dumps({'type': 'warning', 'message': f'Erreur lecture facture {nom}: {e}'})}\n\n"

                step += 1
                yield evt_pbar(pct(), nom)

            if not lignes_factures:
                yield f"data: {json.dumps({'type': 'erreur', 'message': 'Aucune ligne de facture valide trouvée.'})}\n\n"
                yield f"data: {json.dumps({'type': 'done'})}\n\n"
                return

            # ── 3. Rapprochement ──────────────────────────────────────────
            yield evt_pbar(97, "Rapprochement…")
            yield f"data: {json.dumps({'type': 'info', 'message': f'Rapprochement de {len(lignes_factures)} lignes…'})}\n\n"

            resultats = []
            for ligne in lignes_factures:
                plaque = ligne['plaque']
                passages_reel = len(passages_peage.get(plaque, []))
                passages_fact = ligne['passages_factures']
                ecart         = passages_fact - passages_reel
                gares_reel    = list(set(passages_peage.get(plaque, [])))

                resultats.append({
                    'client':            ligne['client'],
                    'plaque':            plaque,
                    'gare_facture':      ligne['gare'],
                    'passages_factures': passages_fact,
                    'passages_peage':    passages_reel,
                    'ecart':             ecart,
                    'montant_facture':   ligne['montant_facture'],
                    'gares_peage':       ', '.join(g for g in gares_reel if g) or '—',
                    'statut': 'ok' if ecart == 0 else ('surfacture' if ecart > 0 else 'sousfacture'),
                })

            totaux = {
                'total_lignes':         len(resultats),
                'total_passages_fact':  sum(r['passages_factures'] for r in resultats),
                'total_passages_peage': sum(r['passages_peage']    for r in resultats),
                'total_ecart':          sum(r['ecart']             for r in resultats),
                'total_montant':        sum(r['montant_facture']   for r in resultats),
                'nb_ok':                sum(1 for r in resultats if r['statut'] == 'ok'),
                'nb_surfacture':        sum(1 for r in resultats if r['statut'] == 'surfacture'),
                'nb_sousfacture':       sum(1 for r in resultats if r['statut'] == 'sousfacture'),
            }

            yield evt_pbar(100, "Terminé")
            yield f"data: {json.dumps({'type': 'resultat', 'data': resultats, 'totaux': totaux})}\n\n"

        except Exception as e:
            traceback.print_exc()
            yield f"data: {json.dumps({'type': 'erreur', 'message': str(e)})}\n\n"

        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, timeout_keep_alive=600)
